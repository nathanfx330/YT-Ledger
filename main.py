# main.py

import configparser
import json
import logging
import subprocess
import sys
import threading
import concurrent.futures
from datetime import datetime
from pathlib import Path
import io
import requests
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from urllib.parse import quote, urlparse, parse_qs
from report_pdf import create_pdf_report
from report_html import create_html_report

# --- Setup Centralized Logging ---
log_file_path = Path("yt_ledger.log")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] - %(message)s",
    handlers=[
        logging.FileHandler(log_file_path, mode='w', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)

def sanitize_filename(name):
    """Removes invalid characters from a string so it can be used as a filename."""
    invalid_chars = set('<>:"/\\|?*')
    return "".join(char for char in name if char not in invalid_chars)[:150]

def stream_handler(stream, log_level_info, log_level_warn):
    """Redirects a stream (like stderr) to the logging system."""
    with stream:
        for line in iter(stream.readline, ''):
            line_stripped = line.strip()
            if line_stripped:
                if line_stripped.lower().startswith(('warning:', 'error:')):
                    log_level_warn(f"[yt-dlp] {line_stripped}")
                else:
                    log_level_info(f"[yt-dlp] {line_stripped}")

def process_playlist_with_yt_dlp(playlist_urls, single_video_ids, video_folder, preferred_resolution, ffmpeg_location, download_videos_flag, archive_file, cookies_file):
    """
    Uses yt-dlp to fetch video information from playlists and single videos, optionally downloading them.
    """
    all_found_ids = set() # Use a set to automatically handle duplicates

    if playlist_urls:
        logging.info(f"Scanning {len(playlist_urls)} playlist(s) to get a list of all available video IDs...")
        for playlist_url in playlist_urls:
            logging.info(f"  -> Scanning playlist: {playlist_url}")
            try:
                command = [
                    'yt-dlp', '--ignore-config', '--no-warnings',
                    '--ignore-errors', '--flat-playlist', '--print', '%(id)s', playlist_url
                ]

                # --- ADD COOKIES TO PLAYLIST SCAN ---
                if cookies_file and cookies_file.exists():
                    logging.info(f"  -> Using cookies from: {cookies_file}")
                    command += ['--cookies', str(cookies_file)]

                process = subprocess.run(command, capture_output=True, text=True, encoding='utf-8', check=True)
                
                playlist_specific_ids = {line for line in process.stdout.splitlines() if line.strip()}
                if not playlist_specific_ids:
                    logging.warning(f"  -> Playlist {playlist_url} is empty or contains no available videos.")
                    continue
                
                logging.info(f"  -> Found {len(playlist_specific_ids)} videos in this playlist.")
                all_found_ids.update(playlist_specific_ids)

            except FileNotFoundError:
                logging.critical("'yt-dlp' command not found. Please ensure yt-dlp is installed and in your system's PATH.")
                return None
            except subprocess.CalledProcessError as e:
                logging.error(f"Could not scan playlist {playlist_url}. It may be private or invalid. Skipping. Error: {e.stderr}")
                continue
    
    if single_video_ids:
        logging.info(f"Adding {len(single_video_ids)} individual video ID(s) to the processing queue.")
        all_found_ids.update(single_video_ids)

    video_ids = sorted(list(all_found_ids))
    total_videos = len(video_ids)

    if total_videos == 0:
        logging.warning("No unique videos were found from the provided playlists or single IDs. Nothing to do.")
        return []
        
    logging.info(f"\n--> Found {total_videos} unique videos across all sources. They will now be processed individually.")

    video_metadata_list = []
    for i, video_id in enumerate(video_ids, 1):
        video_url = f"https://www.youtube.com/watch?v={video_id}"
        logging.info(f"\n--- PROCESSING VIDEO {i} of {total_videos}: {video_url} ---")

        output_template = video_folder / '%(title)s [%(id)s].%(ext)s'
        command = [
            'yt-dlp', '--ignore-config', '--print-json', '--ignore-errors',
        ]

        if archive_file:
            logging.info(f"  -> Archive functionality is ON. Using file: {archive_file}")
            command += ['--download-archive', str(archive_file)]
        else:
            logging.info("  -> Archive functionality is OFF. All videos will be processed.")

        # --- ADD COOKIES TO MAIN DOWNLOAD/METADATA COMMAND ---
        if cookies_file and cookies_file.exists():
            logging.info(f"  -> Using cookies from: {cookies_file}")
            command += ['--cookies', str(cookies_file)]

        if download_videos_flag:
            video_folder.mkdir(parents=True, exist_ok=True)
            command += [
                '-o', str(output_template),
                '-f', f'bestvideo[ext=mp4][height<={preferred_resolution}]+bestaudio[ext=m4a]/best[ext=mp4]/best',
                '--merge-output-format', 'mp4',
                '--write-info-json', '--write-sub', '--write-auto-sub', '--sub-format', 'srt'
            ]
            if ffmpeg_location:
                command += ['--ffmpeg-location', ffmpeg_location]

        command.append(video_url)

        try:
            process = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, encoding='utf-8')
            stderr_thread = threading.Thread(target=stream_handler, args=(process.stderr, logging.info, logging.warning))
            stderr_thread.start()

            stdout_lines = []
            with process.stdout:
                for line in iter(process.stdout.readline, ''):
                    stdout_lines.append(line)

            stderr_thread.join()
            process.wait()

            if not stdout_lines:
                logging.warning(f"  -> FAILED to get metadata for video {video_id}. It may have been skipped by the archive file. Skipping.")
                continue

            video_info = json.loads(stdout_lines[-1])
            video_title = video_info.get('title', 'N/A')
            channel_name = video_info.get('channel', 'N/A')
            upload_date_raw = video_info.get('upload_date')
            upload_date = datetime.strptime(upload_date_raw, '%Y%m%d').strftime('%Y-%m-%d') if upload_date_raw else "N/A"
            video_path = Path(video_info.get('_filename', "Not Downloaded"))

            video_metadata_list.append({
                'id': video_info.get('id'),
                'title': video_title,
                'channel': channel_name,
                'thumbnail_url': video_info.get('thumbnail'),
                'upload_date': upload_date,
                'video_path': video_path,
                'description': video_info.get('description', 'N/A')
            })

            logging.info(f"  -> Successfully retrieved metadata for '{video_title}'.")

            if download_videos_flag and video_path.name != "Not Downloaded":
                clean_metadata = {
                    "video_title": video_title,
                    "channel_name": channel_name,
                    "youtube_url": video_info.get("webpage_url"),
                    "original_upload_date": upload_date,
                    "thumbnail_url": video_info.get('thumbnail'),
                    "local_file_path": video_path.as_posix(),
                    "description": video_info.get('description', 'N/A')
                }
                meta_filepath = video_path.with_suffix('.meta.json')
                with open(meta_filepath, 'w', encoding='utf-8') as f:
                    json.dump(clean_metadata, f, indent=4, ensure_ascii=False)

        except Exception as e:
            logging.critical(f"A critical error occurred while processing video {video_id}. Skipping. Error: {e}")
            continue

    logging.info(f"\n--> Finished processing all available videos. Total collected: {len(video_metadata_list)}.")
    return video_metadata_list

def download_thumbnail(video_data, thumbs_folder, session):
    """Downloads a single thumbnail for a given video."""
    video_id = video_data.get('id')
    video_title = video_data.get('title', 'N/A')
    thumbnail_url = video_data.get('thumbnail_url')

    if not thumbnail_url:
        return video_id, None

    safe_title = sanitize_filename(video_title)
    local_thumb_path = thumbs_folder / f"{safe_title} [{video_id}].jpg"

    if local_thumb_path.exists():
        return video_id, local_thumb_path

    try:
        res = session.get(thumbnail_url, timeout=15)
        res.raise_for_status()
        with open(local_thumb_path, 'wb') as f:
            f.write(res.content)
        return video_id, local_thumb_path
    except requests.exceptions.RequestException as e:
        logging.warning(f"  > Could not download thumbnail for {video_id}. Error: {e}")
        return video_id, None

def download_all_thumbnails_parallel(video_list, thumbs_folder):
    """Downloads all thumbnails in parallel using a thread pool."""
    logging.info("\nDownloading thumbnails...")
    thumbs_folder.mkdir(exist_ok=True)
    thumb_map = {}

    if not video_list:
        logging.info("No videos processed, skipping thumbnail download.")
        return thumb_map

    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
        with requests.Session() as session:
            future_to_video = {
                executor.submit(download_thumbnail, video_data, thumbs_folder, session): video_data
                for video_data in video_list
            }
            processed_thumbs = 0
            for future in concurrent.futures.as_completed(future_to_video):
                processed_thumbs += 1
                video_id, thumb_path = future.result()
                if thumb_path:
                    thumb_map[video_id] = thumb_path
                sys.stdout.write(f"\r  -> Processed {processed_thumbs}/{len(video_list)} thumbnails...")
                sys.stdout.flush()

    sys.stdout.write("\n")
    logging.info("Thumbnail download stage complete.")
    return thumb_map

def create_spreadsheet(video_list, thumb_map, output_filename):
    """Creates an Excel spreadsheet report."""
    if not video_list: return
    wb = Workbook()
    ws = wb.active
    ws.title = "YT-Ledger Report"
    
    headers = ["Thumbnail", "Video Title", "Channel Name", "Full YouTube URL", "Original Upload Date", "Local File Path"]
    ws.append(headers)

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 60 
    ws.column_dimensions['C'].width = 30 
    ws.column_dimensions['D'].width = 50 
    ws.column_dimensions['E'].width = 20 
    ws.column_dimensions['F'].width = 70 

    logging.info("\nBuilding the Excel spreadsheet...")
    for index, video_data in enumerate(video_list, start=2):
        ws.row_dimensions[index].height = 75
        video_id = video_data.get('id')

        if video_id in thumb_map:
            try:
                img_pillow = Image.open(thumb_map[video_id])
                img_pillow.thumbnail((120, 90))
                png_stream = io.BytesIO()
                img_pillow.save(png_stream, format='PNG')
                png_stream.seek(0)
                ws.add_image(OpenpyxlImage(png_stream), f'A{index}')
            except Exception as e:
                logging.warning(f"  > Warning: Could not process thumbnail for {video_id}. Error: {e}")

        video_url = f"https://www.youtube.com/watch?v={video_id}"
        ws[f'B{index}'] = video_data.get('title', 'N/A')
        ws[f'C{index}'] = video_data.get('channel', 'N/A')
        
        cell = ws[f'D{index}']
        cell.value = video_url
        cell.hyperlink = video_url
        cell.style = "Hyperlink"
        
        ws[f'E{index}'] = video_data.get('upload_date', 'N/A')
        
        video_path = video_data.get('video_path')
        ws[f'F{index}'] = video_path.as_posix() if video_path.name != "Not Downloaded" else "Not Downloaded"

    try:
        wb.save(output_filename)
        logging.info(f"✅ Success! Spreadsheet saved as '{output_filename}'")
    except Exception as e:
        logging.error(f"❌ Error saving Excel file: {e}")

def main():
    """Main function to run the archival script."""
    project_name = "YT-Ledger v1.0"

    config = configparser.ConfigParser(interpolation=None)
    config_path = Path('config.ini')
    if not config_path.exists():
        logging.critical("CRITICAL ERROR: config.ini not found. Please create it from the template.")
        return

    try:
        config_lines = config_path.read_text(encoding='utf-8').splitlines()
        config_string = "\n".join(line for line in config_lines if not line.strip().startswith('#') and not line.strip().startswith(';'))
        config.read_string(config_string)
        
        ids_str = config.get('youtube', 'playlist_id', fallback='')
        report_title = config.get('youtube', 'report_title', fallback=f"YouTube Archive Report - {datetime.now().strftime('%Y-%m-%d')}")
        
        download_videos_flag = config.getboolean('downloads', 'download_videos')
        video_folder = Path(config.get('downloads', 'video_folder'))
        preferred_resolution = config.get('downloads', 'preferred_resolution', fallback='1080')
        ffmpeg_location = config.get('downloads', 'ffmpeg_location', fallback=None) or None
        
        archive_file_str = config.get('downloads', 'archive_file', fallback='').strip()
        archive_file = Path(archive_file_str) if archive_file_str else None

        # --- READ COOKIES FILE PATH FROM CONFIG ---
        cookies_file_str = config.get('downloads', 'cookies_file', fallback='').strip()
        cookies_file = Path(cookies_file_str) if cookies_file_str else None
        
        output_xls = Path(config.get('outputs', 'output_file_xls'))
        output_html = Path(config.get('outputs', 'output_file_html'))
        output_pdf = Path(config.get('outputs', 'output_file_pdf'))
        thumbs_folder = Path(config.get('outputs', 'thumbs_folder'))
        template_html = Path(config.get('outputs', 'template_file_html'))
        
    except Exception as e:
        logging.critical(f"CRITICAL ERROR reading config.ini: {e}")
        return

    if not ids_str or 'Please paste playlist ID' in ids_str:
        logging.critical("CRITICAL ERROR: Please set your playlist_id in config.ini")
        return

    all_input_items = [item.strip() for item in ids_str.replace('+', '\n').splitlines() if item.strip()]
    if not all_input_items:
        logging.critical("CRITICAL ERROR: playlist_id field in config.ini is empty or contains only whitespace.")
        return

    playlist_urls = []
    single_video_ids = set()
    playlist_prefixes = ('PL', 'FL', 'UU', 'RD')
    logging.info("Parsing all items from the 'playlist_id' field in config...")

    for item in all_input_items:
        if item.startswith('http://') or item.startswith('https://'):
            try:
                parsed_url = urlparse(item)
                hostname = parsed_url.hostname.lower() if parsed_url.hostname else ''
                
                if 'youtube.com' in hostname:
                    query_params = parse_qs(parsed_url.query)
                    if 'list' in query_params:
                        playlist_urls.append(f"https://www.youtube.com/playlist?list={query_params['list'][0]}")
                    elif 'v' in query_params:
                        single_video_ids.add(query_params['v'][0])
                elif 'youtu.be' in hostname:
                    video_id = parsed_url.path.strip('/')
                    if video_id: single_video_ids.add(video_id)
            except Exception: pass
        
        elif item.upper().startswith(playlist_prefixes):
            playlist_urls.append(f"https://www.youtube.com/playlist?list={item}")
        
        elif len(item) == 11 and not item.isspace():
             single_video_ids.add(item)
    
    logging.info(f"--> Found {len(playlist_urls)} playlist(s) and {len(single_video_ids)} single video(s) to process.")

    video_list = process_playlist_with_yt_dlp(
        playlist_urls, single_video_ids, video_folder, preferred_resolution, 
        ffmpeg_location, download_videos_flag, archive_file, cookies_file
    )
    
    if not video_list:
        logging.warning("\nHalting script because no video data could be processed.")
        return

    video_list.sort(key=lambda v: v.get('upload_date', '0000-00-00'), reverse=True)

    thumb_map = download_all_thumbnails_parallel(video_list, thumbs_folder)
    create_spreadsheet(video_list, thumb_map, output_xls)
    create_html_report(video_list, thumb_map, report_title, project_name, template_html, output_html)
    
    # --- The PDF module now handles its own configuration ---
    create_pdf_report(video_list, thumb_map, output_pdf, report_title, project_name)
    
    logging.info("\n\nAll tasks complete.")

if __name__ == '__main__':
    main()
